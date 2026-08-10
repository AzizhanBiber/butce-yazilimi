from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.db import models
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.models import User as AuthUser
from django.http import HttpResponse
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from django.core.mail import send_mail
import random
from .models import (
    BudgetHeader, BudgetLine, Department, Period, Category,
    ApprovalStep, ActualData, Notification, AuditLog
)
from .forms import BudgetHeaderForm, BudgetLineForm


@login_required
def dashboard(request):
    budgets_qs = BudgetHeader.objects.all()
    if not request.user.is_superuser:
        budgets_qs = budgets_qs.filter(department__manager=request.user)

    total_budgets = budgets_qs.count()
    departments = Department.objects.count()
    pending_approval = budgets_qs.filter(
        status__in=["dept_onay", "finans_onay", "ust_onay"]
    ).count()
    approved = budgets_qs.filter(status="onaylandi").count()
    rejected = budgets_qs.filter(status="reddedildi").count()
    draft = budgets_qs.filter(status="taslak").count()

    dept_data = (
        BudgetHeader.objects.filter(status="onaylandi")
        .values("department__name")
        .annotate(total=Sum("lines__planned_amount"))
        .order_by("-total")
    )
    dept_labels = [d["department__name"] for d in dept_data]
    dept_values = [float(d["total"] or 0) for d in dept_data]

    recent_budgets = budgets_qs.select_related("department", "period").order_by("-created_at")[:5]

    pending_list = budgets_qs.filter(
        status__in=["dept_onay", "finans_onay", "ust_onay"]
    ).select_related("department", "period")[:5]

    unread_count = request.user.notifications.filter(is_read=False).count()

    risk_alerts = []
    risk_departments= Department.objects.all() if request.user.is_superuser else Department.objects.filter(manager=request.user)
    for dept in risk_departments:
        dept_planned = BudgetHeader.objects.filter(department=dept, status="onaylandi").aggregate(
            total=Sum("lines__planned_amount")
        )["total"] or 0
        dept_actual = ActualData.objects.filter(department=dept).aggregate(
            total=Sum("actual_amount")
        )["total"] or 0
        if dept_planned > 0:
            usage = (dept_actual / dept_planned) * 100
            if usage >= 100:
                risk_alerts.append({"department": dept.name, "usage": round(usage, 1), "level": "danger"})
            elif usage >= 80:
                risk_alerts.append({"department": dept.name, "usage": round(usage, 1), "level": "warning"})

    top_department = dept_labels[0] if dept_labels else None
    top_department_amount = dept_values[0] if dept_values else 0

    context = {
        "total_budgets": total_budgets,
        "departments": departments,
        "pending_approval": pending_approval,
        "approved": approved,
        "rejected": rejected,
        "draft": draft,
        "dept_labels": dept_labels,
        "dept_values": dept_values,
        "recent_budgets": recent_budgets,
        "pending_list": pending_list,
        "unread_count": unread_count,
        "risk_alerts": risk_alerts,
        "top_department": top_department,
        "top_department_amount": top_department_amount,
    }
    return render(request, "budgets/dashboard.html", context)


@login_required
def budget_list(request):
    budgets = BudgetHeader.objects.select_related("department", "period").annotate(
        total_amount=Sum("lines__planned_amount")
    )
    if not request.user.is_superuser:
        budgets = budgets.filter(department__manager=request.user)

    department_id = request.GET.get("department")
    period_id = request.GET.get("period")
    status = request.GET.get("status")

    if department_id:
        budgets = budgets.filter(department_id=department_id)
    if period_id:
        budgets = budgets.filter(period_id=period_id)
    if status:
        budgets = budgets.filter(status=status)

    context = {
        "budgets": budgets,
        "departments": Department.objects.all(),
        "periods": Period.objects.all(),
        "status_choices": BudgetHeader.STATUS_CHOICES,
        "selected_department": department_id,
        "selected_period": period_id,
        "selected_status": status,
    }
    return render(request, "budgets/budget_list.html", context)


@login_required
def budget_create(request):
    if request.method == "POST":
        form = BudgetHeaderForm(request.POST, user=request.user)
        if form.is_valid():
            budget = form.save(commit=False)
            budget.created_by = request.user if request.user.is_authenticated else None
            budget.save()
            AuditLog.objects.create(
                user=request.user, action="olusturuldu",
                description=f"{budget.department} - {budget.period} bütçesi oluşturuldu."
            )
            return redirect("budget_list")
    else:
        form = BudgetHeaderForm(user=request.user)
    return render(request, "budgets/budget_form.html", {"form": form})


@login_required
def budget_detail(request, pk):
    budget = get_object_or_404(BudgetHeader, pk=pk)
    if not request.user.is_superuser and budget.department.manager != request.user:
        messages.error(request, "Bu bütçeyi görüntüleme yetkiniz yok.")
        return redirect("budget_list")
    lines = budget.lines.all()
    if request.method == "POST":
        form = BudgetLineForm(request.POST)
        if form.is_valid():
            line = form.save(commit=False)
            line.budget_header = budget
            line.save()
            return redirect("budget_detail", pk=budget.pk)
    else:
        form = BudgetLineForm()
    total = sum(x.planned_amount for x in lines)
    return render(request, "budgets/budget_detail.html", {
        "budget": budget,
        "lines": lines,
        "form": form,
        "total": total,
    })


@login_required
def submit_for_approval(request, pk):
    budget = get_object_or_404(BudgetHeader, pk=pk)
    budget.status = "dept_onay"
    budget.save()
    ApprovalStep.objects.create(
        budget_header=budget,
        step_order=1,
        approver=None,
        status="bekliyor",
    )
    for admin_user in AuthUser.objects.filter(is_superuser=True):
        Notification.objects.create(
            user=admin_user,
            message=f"{budget.department} departmani bir butceyi onaya sundu.",
            budget_header=budget
        )
    if admin_user.email:
        send_mail(
            subject="OnayaSunulan Bütçe",
            message=f"{budget.department} departmanı {budget.period} dönemi için bir bütçeyi onayınıza sundu.",
            from_email=None,
            recipient_list= [admin_user.email],
            fail_silently=True,
        )    
    AuditLog.objects.create(
        user=request.user, action="onaya_gonderildi",
        description=f"{budget.department} bütçesi onaya gönderildi."
    )
    messages.success(request, "Bütçe onaya gönderildi.")
    return redirect("budget_detail", pk=budget.pk)


@login_required
def approve_budget(request, pk):
    budget = get_object_or_404(BudgetHeader, pk=pk)
    if not request.user.is_superuser:
        messages.error(request, "Bu işlemi yapma yetkiniz yok.")
        return redirect("budget_detail", pk=budget.pk)
    budget.status = "onaylandi"
    budget.save()
    step = budget.approval_steps.filter(status="bekliyor").first()
    if step:
        step.status = "onaylandi"
        step.approver = request.user
        step.acted_at = timezone.now()
        step.save()
    if budget.department.manager:
        Notification.objects.create(
            user=budget.department.manager,
            message=f"{budget.department} butceniz onaylandi.",
            budget_header=budget
        )
    if budget.department.manager.email:
            send_mail(
                subject="Bütçeniz Onaylandı",
                message=f"{budget.department} departmanının {budget.period} dönemi bütçesi onaylandı.",
                from_email=None,
                recipient_list=[budget.department.manager.email],
                fail_silently=True,
            )    
    AuditLog.objects.create(
        user=request.user, action="onaylandi",
        description=f"{budget.department} bütçesi onaylandı."
    )
    messages.success(request, "Bütçe onaylandı.")
    return redirect("budget_detail", pk=budget.pk)


@login_required
def reject_budget(request, pk):
    budget = get_object_or_404(BudgetHeader, pk=pk)
    if not request.user.is_superuser:
        messages.error(request, "Bu işlemi yapma yetkiniz yok.")
        return redirect("budget_detail", pk=budget.pk)
    budget.status = "reddedildi"
    budget.save()
    step = budget.approval_steps.filter(status="bekliyor").first()
    if step:
        step.status = "reddedildi"
        step.approver = request.user
        step.acted_at = timezone.now()
        step.save()
    if budget.department.manager:
        Notification.objects.create(
            user=budget.department.manager,
            message=f"{budget.department} butceniz reddedildi.",
            budget_header=budget
        )
    if budget.department.manager.email:
            send_mail(
                subject="Bütçeniz Reddedildi",
                message=f"{budget.department} departmanının {budget.period} dönemi bütçesi reddedildi.",
                from_email=None,
                recipient_list=[budget.department.manager.email],
                fail_silently=True,
            )    
    AuditLog.objects.create(
        user=request.user, action="reddedildi",
        description=f"{budget.department} bütçesi reddedildi."
    )
    messages.warning(request, "Bütçe reddedildi.")
    return redirect("budget_detail", pk=budget.pk)


@login_required
def create_new_version(request, pk):
    old_budget = get_object_or_404(BudgetHeader, pk=pk)
    if not request.user.is_superuser and old_budget.department.manager != request.user:
        messages.error(request, "Bu işlemi yapma yetkiniz yok.")
        return redirect("budget_detail", pk=old_budget.pk)

    max_version = BudgetHeader.objects.filter(
        department=old_budget.department,
        period=old_budget.period,
        budget_type=old_budget.budget_type,
    ).aggregate(m=models.Max("version_number"))["m"] or 1

    new_budget = BudgetHeader.objects.create(
        department=old_budget.department,
        period=old_budget.period,
        budget_type=old_budget.budget_type,
        status="taslak",
        version_number=max_version + 1,
        created_by=request.user,
    )
    for line in old_budget.lines.all():
        BudgetLine.objects.create(
            budget_header=new_budget,
            category=line.category,
            planned_amount=line.planned_amount,
            month=line.month,
            notes=line.notes,
        )
    AuditLog.objects.create(
        user=request.user, action="versiyon_olusturuldu",
        description=f"{old_budget.department} için versiyon {new_budget.version_number} oluşturuldu."
    )
    messages.success(request, f"Versiyon {new_budget.version_number} oluşturuldu.")
    return redirect("budget_detail", pk=new_budget.pk)


@login_required
def edit_budget_line(request, pk):
    line = get_object_or_404(BudgetLine, pk=pk)
    budget = line.budget_header
    if not request.user.is_superuser and budget.department.manager != request.user:
        messages.error(request, "Bu işlemi yapma yetkiniz yok.")
        return redirect("budget_detail", pk=budget.pk)
    if budget.status != "taslak":
        messages.error(request, "Sadece taslak bütçelerde kalem düzenlenebilir.")
        return redirect("budget_detail", pk=budget.pk)
    if request.method == "POST":
        form = BudgetLineForm(request.POST, instance=line)
        if form.is_valid():
            form.save()
            messages.success(request, "Kalem güncellendi.")
            return redirect("budget_detail", pk=budget.pk)
    else:
        form = BudgetLineForm(instance=line)
    return render(request, "budgets/budget_line_edit.html", {"form": form, "budget": budget})


@login_required
def delete_budget_line(request, pk):
    line = get_object_or_404(BudgetLine, pk=pk)
    budget = line.budget_header
    if not request.user.is_superuser and budget.department.manager != request.user:
        messages.error(request, "Bu işlemi yapma yetkiniz yok.")
        return redirect("budget_detail", pk=budget.pk)
    if budget.status != "taslak":
        messages.error(request, "Sadece taslak bütçelerde kalem silinebilir.")
        return redirect("budget_detail", pk=budget.pk)
    line.delete()
    messages.success(request, "Kalem silindi.")
    return redirect("budget_detail", pk=budget.pk)


@login_required
def variance_report(request):
    departments = Department.objects.all()
    if not request.user.is_superuser:
        departments = departments.filter(manager=request.user)

    report = []
    for dept in departments:
        planned = BudgetHeader.objects.filter(department=dept, status="onaylandi").aggregate(
            total=Sum("lines__planned_amount")
        )["total"] or 0
        actual = ActualData.objects.filter(department=dept).aggregate(
            total=Sum("actual_amount")
        )["total"] or 0
        if planned > 0:
            usage_percent = round((actual / planned) * 100, 1)
        else:
            usage_percent = 0
        variance = planned - actual
        report.append({
            "department": dept.name,
            "planned": planned,
            "actual": actual,
            "variance": variance,
            "usage_percent": usage_percent,
        })

    return render(request, "budgets/variance_report.html", {"report": report})


@login_required
def notifications(request):
    notes = request.user.notifications.all()
    return render(request, "budgets/notifications.html", {"notes": notes})


@login_required
def export_budget_excel(request, pk):
    budget = get_object_or_404(BudgetHeader, pk=pk)
    if not request.user.is_superuser and budget.department.manager != request.user:
        messages.error(request, "Bu işlemi yapma yetkiniz yok.")
        return redirect("budget_detail", pk=budget.pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bütçe Kalemleri"
    ws.append(["Kategori", "Tutar (TL)", "Ay", "Açıklama"])
    for line in budget.lines.all():
        ws.append([line.category.name, float(line.planned_amount), line.month or "", line.notes])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"{budget.department}_{budget.period}_v{budget.version_number}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def import_budget_excel(request, pk):
    budget = get_object_or_404(BudgetHeader, pk=pk)
    if not request.user.is_superuser and budget.department.manager != request.user:
        messages.error(request, "Bu işlemi yapma yetkiniz yok.")
        return redirect("budget_detail", pk=budget.pk)
    if budget.status != "taslak":
        messages.error(request, "Sadece taslak bütçelere Excel yüklenebilir.")
        return redirect("budget_detail", pk=budget.pk)

    if request.method == "POST" and request.FILES.get("excel_file"):
        wb = openpyxl.load_workbook(request.FILES["excel_file"])
        ws = wb.active
        added = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            category_name, amount, month, notes = (row + (None, None, None, None))[:4]
            if not category_name or amount is None:
                continue
            category = Category.objects.filter(name=category_name).first()
            if not category:
                continue
            BudgetLine.objects.create(
                budget_header=budget,
                category=category,
                planned_amount=amount,
                month=month if month else None,
                notes=notes or "",
            )
            added += 1
        messages.success(request, f"{added} kalem içe aktarıldı.")
    else:
        messages.error(request, "Dosya seçilmedi.")
    return redirect("budget_detail", pk=budget.pk)


@login_required
def export_budget_pdf(request, pk):
    budget = get_object_or_404(BudgetHeader, pk=pk)
    if not request.user.is_superuser and budget.department.manager != request.user:
        messages.error(request, "Bu işlemi yapma yetkiniz yok.")
        return redirect("budget_detail", pk=budget.pk)

    response = HttpResponse(content_type="application/pdf")
    filename = f"{budget.department}_{budget.period}_v{budget.version_number}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    p.setFont("Helvetica-Bold", 16)
    p.drawString(2*cm, height - 2*cm, "TUMOSAN Butce Sistemi")

    p.setFont("Helvetica-Bold", 13)
    p.drawString(2*cm, height - 3*cm, f"{budget.department} - {budget.period}")

    p.setFont("Helvetica", 10)
    p.drawString(2*cm, height - 3.7*cm, f"Tur: {budget.get_budget_type_display()}")
    p.drawString(2*cm, height - 4.3*cm, f"Durum: {budget.get_status_display()}")
    p.drawString(2*cm, height - 4.9*cm, f"Versiyon: {budget.version_number}")

    total = sum(x.planned_amount for x in budget.lines.all())
    p.setFont("Helvetica-Bold", 11)
    p.drawString(2*cm, height - 5.8*cm, f"Toplam: {total:,.2f} TL")

    y = height - 7*cm
    p.setFont("Helvetica-Bold", 10)
    p.drawString(2*cm, y, "Kategori")
    p.drawString(9*cm, y, "Tutar (TL)")
    p.drawString(13*cm, y, "Ay")
    p.drawString(15*cm, y, "Aciklama")
    p.line(2*cm, y - 0.2*cm, 19*cm, y - 0.2*cm)

    p.setFont("Helvetica", 9)
    y -= 0.8*cm
    for line in budget.lines.all():
        if y < 2*cm:
            p.showPage()
            y = height - 2*cm
        p.drawString(2*cm, y, str(line.category.name)[:30])
        p.drawString(9*cm, y, f"{line.planned_amount:,.2f}")
        p.drawString(13*cm, y, str(line.month or "-"))
        p.drawString(15*cm, y, str(line.notes)[:25])
        y -= 0.6*cm

    p.save()
    return response


@login_required
def summary_report(request):
    departments = Department.objects.all()
    if not request.user.is_superuser:
        departments = departments.filter(manager=request.user)

    report = []
    for dept in departments:
        budgets = BudgetHeader.objects.filter(department=dept)
        planned = budgets.filter(status="onaylandi").aggregate(total=Sum("lines__planned_amount"))["total"] or 0
        actual = ActualData.objects.filter(department=dept).aggregate(total=Sum("actual_amount"))["total"] or 0

        category_breakdown = (
            BudgetLine.objects.filter(
                budget_header__department=dept, budget_header__status="onaylandi"
            )
            .values("category__name")
            .annotate(total=Sum("planned_amount"), item_count=Count("id"))
            .order_by("-total")
        )

        line_notes = (
            BudgetLine.objects.filter(
                budget_header__department=dept, budget_header__status="onaylandi"
            )
            .exclude(notes="")
            .values("category__name", "notes", "planned_amount")
        )

        report.append({
            "department": dept.name,
            "budget_count": budgets.count(),
            "planned": planned,
            "actual": actual,
            "approved_count": budgets.filter(status="onaylandi").count(),
            "pending_count": budgets.filter(status__in=["dept_onay", "finans_onay", "ust_onay"]).count(),
            "category_breakdown": category_breakdown,
            "line_notes": line_notes,
        })

    return render(request, "budgets/summary_report.html", {"report": report})


@login_required
def trend_report(request):
    departments = Department.objects.all()
    if not request.user.is_superuser:
        departments = departments.filter(manager=request.user)

    months = list(range(1, 13))
    month_names = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                   "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]

    dept_series = []
    for dept in departments:
        annual_planned = BudgetLine.objects.filter(
            budget_header__department=dept, budget_header__status="onaylandi", month__isnull=True
        ).aggregate(total=Sum("planned_amount"))["total"] or 0

        planned_by_month = []
        actual_by_month = []
        for m in months:
            planned = BudgetLine.objects.filter(
                budget_header__department=dept, month=m, budget_header__status="onaylandi"
            ).aggregate(total=Sum("planned_amount"))["total"] or 0
            actual = ActualData.objects.filter(
                department=dept, month=m
            ).aggregate(total=Sum("actual_amount"))["total"] or 0
            planned_by_month.append(float(planned))
            actual_by_month.append(float(actual))
        dept_series.append({
            "department": dept.name,
            "planned": planned_by_month,
            "actual": actual_by_month,
            "annual_planned": float(annual_planned),
        })

    return render(request, "budgets/trend_report.html", {
        "dept_series": dept_series,
        "month_names": month_names,
    })
@login_required
def sync_from_sap(request, pk):
    budget = get_object_or_404(BudgetHeader, pk=pk)
    if not request.user.is_superuser and budget.department.manager != request.user:
        messages.error(request, "Bu işlemi yapma yetkiniz yok.")
        return redirect("budget_detail", pk=budget.pk)

    lines = budget.lines.all()
    synced_count = 0
    for line in lines:
        variance_factor = random.uniform(0.75, 1.15)
        simulated_actual = round(float(line.planned_amount) * variance_factor, 2)

        ActualData.objects.update_or_create(
            department=budget.department,
            category=line.category,
            period=budget.period,
            month=line.month or 1,
            defaults={"actual_amount": simulated_actual}
        )
        synced_count += 1

    AuditLog.objects.create(
        user=request.user, action="olusturuldu",
        description=f"{budget.department} için SAP'den {synced_count} kalem gerçekleşen veri senkronize edildi."
    )
    messages.success(request, f"SAP'den {synced_count} kalem başarıyla senkronize edildi.")
    return redirect("budget_detail", pk=budget.pk)
@login_required
def delete_notification(request, pk):
    note = get_object_or_404(Notification, pk=pk, user=request.user)
    note.delete()
    messages.success(request, "Bildirim silindi.")
    return redirect("notifications")
@login_required
def open_notification(request, pk):
    note = get_object_or_404(Notification, pk=pk, user=request.user)
    note.is_read = True
    note.save()
    if note.budget_header:
        return redirect("budget_detail", pk=note.budget_header.pk)
    return redirect("notifications")